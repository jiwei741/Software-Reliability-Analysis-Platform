from __future__ import annotations

import csv
import io
import os
import re
import json
import threading
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Sequence
from urllib.parse import urlparse
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from flask import Flask, Response, jsonify, render_template, request, send_from_directory

from services.data_store import store
from services.deepseek import DeepSeekError, infer_mapping_from_sample
from services.reliability import (
    SECTION_CHARTS,
    build_dashboard_payload,
    build_section_payload,
)

BASE_DIR = Path(__file__).resolve().parent
DEBUG_MODE = True
ALLOWED_IDENTIFIER = re.compile(r"^[A-Za-z0-9_]+$")
WORKER_ENDPOINT = os.environ.get("WORKER_ENDPOINT")
WORKER_AUTH_TOKEN = os.environ.get("WORKER_AUTH_TOKEN")


def _load_env_file() -> None:
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())


_load_env_file()

app = Flask(
    __name__,
    static_folder="assets",
    static_url_path="/assets",
    template_folder="templates",
)
app.config.update(TEMPLATES_AUTO_RELOAD=True)


def _current_payload() -> Dict[str, object]:
    return build_dashboard_payload(store.get_records())


def _render_export_html(payload: Dict[str, object]) -> str:
    """Render a lightweight standalone HTML with charts/formulas/results for download."""
    data_json = json.dumps(payload, ensure_ascii=False)
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <title>可靠性分析导出</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <script>
    window.MathJax = {{
      tex: {{
        inlineMath: [['\\\\(', '\\\\)'], ['$', '$']],
        displayMath: [['\\\\[', '\\\\]'], ['$$', '$$']]
      }},
      svg: {{ fontCache: 'global' }}
    }};
  </script>
  <script id="MathJax-script" async src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-svg.js"></script>
  <style>
    body {{ font-family: 'Inter', system-ui, -apple-system, sans-serif; margin: 24px; background:#0f172a; color:#e2e8f0; }}
    h1,h2,h3 {{ margin: 0 0 12px; }}
    .grid {{ display:grid; gap:16px; grid-template-columns: repeat(auto-fit,minmax(320px,1fr)); }}
    .card {{ background:#111827; border:1px solid #1f2937; border-radius:12px; padding:16px; box-shadow:0 10px 30px rgba(0,0,0,0.25); }}
    canvas {{ background:#0b1221; }}
    table {{ width:100%; border-collapse:collapse; margin-top:8px; }}
    th,td {{ border-bottom:1px solid #1f2937; padding:6px 8px; text-align:left; }}
    code {{ background:#1f2937; padding:2px 6px; border-radius:6px; }}
  </style>
</head>
<body>
  <h1>可靠性分析导出</h1>
  <p>生成时间：{datetime.utcnow().isoformat()}</p>

  <section>
    <h2>模型参数</h2>
    <div class="grid">
      {''.join(f"<div class='card'><h3>{p['name']}</h3><p>ID: {p['id']}</p><p>α={p['alpha']} β={p['beta']}</p><p>备注: {p['extra']}</p><p>更新时间: {p['updated']}</p></div>" for p in payload.get('model_params', []))}
    </div>
  </section>

  <section>
    <h2>最近记录</h2>
    <div class="card">
      <table>
        <thead><tr><th>模块</th><th>失效</th><th>MTBF</th><th>时长</th><th>时间</th></tr></thead>
        <tbody>
        {''.join(f"<tr><td>{r['module']}</td><td>{r['failures']}</td><td>{r['mtbf']}</td><td>{r['runtime']}</td><td>{r['timestamp']}</td></tr>" for r in payload.get('recent_records', []))}
        </tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>公式</h2>
    <div class="grid" id="formula-container"></div>
  </section>

  <section>
    <h2>图表</h2>
    <div class="grid" id="chart-container"></div>
  </section>

  <script>
    const payload = {data_json};
    const chartContainer = document.getElementById('chart-container');
    (payload.charts || []).forEach((chart) => {{
      const card = document.createElement('div');
      card.className = 'card';
      const title = document.createElement('h3');
      title.textContent = chart.id;
      const canvas = document.createElement('canvas');
      card.appendChild(title);
      card.appendChild(canvas);
      chartContainer.appendChild(card);
      new Chart(canvas.getContext('2d'), {{
        type: chart.type,
        data: {{ labels: chart.labels, datasets: chart.datasets }},
        options: chart.options || {{}}
      }});
    }});

    const formulaContainer = document.getElementById('formula-container');
    const formulaLists = Object.values(payload.formulas || {{}});
    const allFormulas = [];
    formulaLists.forEach((lst) => Array.isArray(lst) && allFormulas.push(...lst));
    allFormulas.forEach((f) => {{
      const card = document.createElement('div');
      card.className = 'card';
      card.innerHTML = `<h3>${{f.title}}</h3><p class="formula-latex">\\\\(${{
        f.latex
      }}\\\\)</p><p>${{f.description || ''}}</p>`;
      formulaContainer.appendChild(card);
    }});

    if (window.MathJax && window.MathJax.typesetPromise) {{
      window.MathJax.typesetPromise([formulaContainer]);
    }}
  </script>
</body>
</html>
"""


def _to_int(value, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _to_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _clean_mapping(raw: Dict[str, str] | None) -> Dict[str, str]:
    if not raw:
        return {}
    cleaned: Dict[str, str] = {}
    for key, value in raw.items():
        if value:
            cleaned[key] = value.strip().lower()
    return cleaned


def _merge_mapping(base: Dict[str, str], override: Dict[str, str] | None) -> Dict[str, str]:
    merged = dict(base)
    if not override:
        return merged
    for key, value in override.items():
        if value:
            merged[key] = value.strip().lower()
    return merged


def _standardize_row(
    row: Dict[str, object],
    *,
    fallback_module: str | None = None,
    mapping: Dict[str, str] | None = None,
) -> Dict[str, object]:
    normalized = {str(key).strip().lower(): value for key, value in row.items() if key is not None}
    mapping = mapping or {}

    def pick(*keys):
        for key in keys:
            mapped = mapping.get(key)
            if mapped and mapped in normalized and normalized[mapped] not in ("", None):
                return normalized[mapped]
        for key in keys:
            if key in normalized and normalized[key] not in ("", None):
                return normalized[key]
        return None

    module = pick("module", "component", "service", "name") or fallback_module or "imported-module"
    failures = _to_int(pick("failures", "failure", "defects", "count") or 0, 0)
    mtbf = _to_float(pick("mtbf", "mean_time_between_failure", "avg_mtbf") or 1.0, 1.0)
    runtime = _to_float(pick("runtime", "runtime_hours", "hours", "executed_hours") or mtbf, mtbf)
    timestamp = pick("timestamp", "time", "created_at", "updated_at", "ts")
    parsed_ts = None
    if timestamp:
        try:
            if isinstance(timestamp, (int, float)):
                parsed_ts = datetime.fromtimestamp(float(timestamp))
            elif isinstance(timestamp, datetime):
                parsed_ts = timestamp
            elif isinstance(timestamp, str):
                parsed_ts = datetime.fromisoformat(timestamp)
        except Exception:
            parsed_ts = None
    if not parsed_ts:
        parsed_ts = datetime.utcnow()
    if parsed_ts.year < datetime.utcnow().year:
        parsed_ts = datetime.utcnow()

    return {
        "module": module,
        "failures": failures,
        "mtbf": mtbf,
        "runtime": runtime,
        "timestamp": parsed_ts.isoformat(),
    }


def _standardize_dataset(rows: Sequence[Dict[str, object]], mapping: Dict[str, str], fallback_module: str | None = None) -> List[Dict[str, object]]:
    return [_standardize_row(row, fallback_module=fallback_module, mapping=mapping) for row in rows]


def _read_csv_raw(file_storage) -> tuple[List[Dict[str, object]], List[str]]:
    text = file_storage.read().decode("utf-8-sig")
    file_storage.stream.close()
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        if not row:
            continue
        cleaned: Dict[str, object] = {}
        for key, value in row.items():
            if key is None:
                continue
            cleaned[str(key)] = value
        rows.append(cleaned)
    headers = reader.fieldnames or (list(rows[0].keys()) if rows else [])
    return rows, headers


def _read_excel_raw(file_storage) -> tuple[List[Dict[str, object]], List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("需要 openpyxl 支持 Excel 导入，请先 `pip install openpyxl`." ) from exc

    data = io.BytesIO(file_storage.read())
    wb = load_workbook(data, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(cell).strip() if cell is not None else f"col_{idx}" for idx, cell in enumerate(rows[0])]
    payload: List[Dict[str, object]] = []
    for raw_row in rows[1:]:
        payload.append({headers[idx]: raw_row[idx] for idx in range(len(headers))})
    return payload, headers


def _load_file_rows(uploaded):
    filename = uploaded.filename.lower()
    if filename.endswith('.csv'):
        return _read_csv_raw(uploaded)
    if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
        return _read_excel_raw(uploaded)
    raise ValueError('仅支持 CSV 或 Excel 文件')


def _parse_mysql_connection(conn_str: str) -> Dict[str, object]:
    parsed = urlparse(conn_str)
    if parsed.scheme not in {"mysql", "mysql+mysqlconnector"}:
        raise ValueError("连接串需以 mysql:// 开头")
    if not parsed.hostname or not parsed.path:
        raise ValueError("连接串缺少 host 或 database")
    return {
        "user": parsed.username or "",
        "password": parsed.password or "",
        "host": parsed.hostname,
        "port": parsed.port or 3306,
        "database": parsed.path.lstrip("/"),
    }


def _fetch_mysql_rows(params: Dict[str, object], table: str, increment: str | None, limit: int) -> List[Dict[str, object]]:
    try:
        import mysql.connector  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("需要 mysql-connector-python 支持 MySQL 导入，请先安装") from exc

    conn = mysql.connector.connect(
        host=params["host"],
        port=params["port"],
        user=params["user"],
        password=params["password"],
        database=params["database"],
    )
    cursor = conn.cursor(dictionary=True)
    order_clause = f" ORDER BY `{increment}` DESC" if increment else ""
    query = f"SELECT * FROM `{table}`{order_clause} LIMIT %s"
    cursor.execute(query, (limit,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def _serialize_rows(rows: Sequence[Dict[str, object]]) -> List[Dict[str, object]]:
    serialized: List[Dict[str, object]] = []
    for row in rows:
        cleaned: Dict[str, object] = {}
        for key, value in row.items():
            if isinstance(value, (str, int, float, bool)) or value is None:
                cleaned[key] = value
            else:
                cleaned[key] = str(value)
        serialized.append(cleaned)
    return serialized


def _persist_user_to_worker(profile: Dict[str, str]) -> Dict[str, object]:
    if not WORKER_ENDPOINT:
        raise RuntimeError("未配置 WORKER_ENDPOINT，无法同步云端")
    payload = json.dumps(profile).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if WORKER_AUTH_TOKEN:
        headers["Authorization"] = f"Bearer {WORKER_AUTH_TOKEN}"
    req = Request(WORKER_ENDPOINT, data=payload, headers=headers, method="POST")
    try:
        with urlopen(req, timeout=4) as resp:
            status = resp.getcode()
            body = resp.read()
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"云端返回 {exc.code}: {detail or exc.reason}") from exc
    except URLError as exc:
        raise RuntimeError(f"无法连接云端: {exc.reason}") from exc
    if status >= 300:
        raise RuntimeError(f"云端返回状态 {status}")
    parsed: Dict[str, object] | None = None
    try:
        parsed = json.loads(body.decode("utf-8") or "{}")
    except Exception:
        parsed = None
    return {"status": status, "data": parsed}


def _fetch_users_from_worker() -> List[Dict[str, object]]:
    if not WORKER_ENDPOINT:
        raise RuntimeError("未配置 WORKER_ENDPOINT，无法同步云端")
    headers = {"Accept": "application/json"}
    if WORKER_AUTH_TOKEN:
        headers["Authorization"] = f"Bearer {WORKER_AUTH_TOKEN}"
    req = Request(WORKER_ENDPOINT, headers=headers, method="GET")
    try:
        with urlopen(req, timeout=4) as resp:
            status = resp.getcode()
            body = resp.read()
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"云端返回 {exc.code}: {detail or exc.reason}") from exc
    except URLError as exc:
        raise RuntimeError(f"无法连接云端: {exc.reason}") from exc
    if status >= 300:
        raise RuntimeError(f"云端返回状态 {status}")
    try:
        parsed = json.loads(body.decode("utf-8") or "{}")
    except Exception:
        parsed = {}
    data = parsed.get("data") if isinstance(parsed, dict) else None
    return data if isinstance(data, list) else []


@app.get("/")
def index() -> str:
    payload = _current_payload()
    return render_template("index.html", dashboard_payload=payload)


@app.get("/api/dashboard")
def api_dashboard() -> Response:
    return jsonify(_current_payload())


@app.get("/api/analyze/<section>")
def analyze_section(section: str) -> Response:
    slug = section.lower()
    if slug not in SECTION_CHARTS:
        return jsonify({"status": "error", "message": "未知的分析板块"}), 404
    payload = build_section_payload(store.get_records(), slug)
    return jsonify({"status": "success", **payload})


@app.get("/api/export/html")
def api_export_html() -> Response:
    html = _render_export_html(_current_payload())
    return Response(
        html,
        mimetype="text/html",
        headers={"Content-Disposition": "attachment; filename=reliability_export.html"},
    )


@app.post("/api/import/manual")
def import_manual() -> Response:
    data = request.get_json(silent=True) or {}
    module = data.get("module")
    if not module:
        return jsonify({"status": "error", "message": "模块名称不能为空"}), 400
    failures = _to_int(data.get("failures"))
    mtbf = _to_float(data.get("mtbf"), 1.0)
    runtime = _to_float(data.get("runtime"), mtbf)
    store.add_record(
        {"module": module, "failures": failures, "mtbf": mtbf, "runtime": runtime},
        source="manual",
    )
    return jsonify({"status": "success", "message": "手动导入成功", "payload": _current_payload()})


@app.post("/api/deepseek/mapping")
def deepseek_mapping() -> Response:
    payload = request.get_json(silent=True) or {}
    headers = payload.get("headers") or []
    rows = payload.get("rows") or []
    if not isinstance(headers, list) or not headers:
        return jsonify({"status": "error", "message": "缺少表头"}), 400
    if not isinstance(rows, list) or not rows:
        return jsonify({"status": "error", "message": "缺少样本数据"}), 400
    sample_rows = rows[:10]
    try:
        mapping = infer_mapping_from_sample(headers, sample_rows, logger=app.logger)
    except DeepSeekError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "mapping": mapping})


@app.post("/api/import/file/sample")
def import_file_sample() -> Response:
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"status": "error", "message": "请先选择 CSV 或 Excel 文件"}), 400
    try:
        rows, headers = _load_file_rows(uploaded)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500
    if not rows:
        return jsonify({"status": "error", "message": "文件内容为空或无法解析"}), 400
    sample = rows[: min(len(rows), 20)]
    return jsonify({"status": "success", "headers": headers, "rows": _serialize_rows(sample)})


@app.post("/api/import/file")
def import_file() -> Response:
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"status": "error", "message": "请先选择 CSV 或 Excel 文件"}), 400

    filename = uploaded.filename.lower()
    tag = request.form.get("tag") or None
    analysis_mode = (request.form.get("analysis-mode") or "preprocess").lower()
    mapping = _clean_mapping(
        {
            "module": request.form.get("map-module"),
            "failures": request.form.get("map-failures"),
            "mtbf": request.form.get("map-mtbf"),
            "runtime": request.form.get("map-runtime"),
        }
    )

    try:
        raw_rows, headers = _load_file_rows(uploaded)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500

    if not raw_rows:
        return jsonify({"status": "error", "message": "文件内容为空或无法解析"}), 400

    if analysis_mode == "deepseek":
        try:
            ai_mapping = infer_mapping_from_sample(headers, raw_rows[:5], logger=app.logger)
            mapping = _merge_mapping(mapping, ai_mapping)
        except DeepSeekError as exc:
            app.logger.warning("DeepSeek 映射失败（文件）：%s", exc)

    normalized_rows = _standardize_dataset(raw_rows, mapping, fallback_module=tag)
    count = store.add_records(normalized_rows, source="file")
    return jsonify({"status": "success", "message": f"文件导入 {count} 条", "payload": _current_payload()})


@app.post("/api/import/mysql/sample")
def import_mysql_sample() -> Response:
    data = request.get_json(silent=True) or {}
    connection = data.get("connection")
    table = data.get("table")
    increment = data.get("increment") or None
    limit = _to_int(data.get("limit") or 50, 50)

    if not connection or not table:
        return jsonify({"status": "error", "message": "连接串和数据表名称不能为空"}), 400
    if not ALLOWED_IDENTIFIER.match(table):
        return jsonify({"status": "error", "message": "仅允许字母、数字、下划线作为表名"}), 400
    if increment and not ALLOWED_IDENTIFIER.match(increment):
        return jsonify({"status": "error", "message": "增量字段仅允许字母、数字、下划线"}), 400

    try:
        params = _parse_mysql_connection(connection)
        rows_raw = _fetch_mysql_rows(params, table, increment, min(limit, 200))
    except (RuntimeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    if not rows_raw:
        return jsonify({"status": "error", "message": "未在数据库中读取到有效数据"}), 400

    headers = list(rows_raw[0].keys())
    sample_rows = rows_raw[: min(len(rows_raw), 10)]
    return jsonify({"status": "success", "headers": headers, "rows": _serialize_rows(sample_rows)})


@app.post("/api/import/mysql")
def import_mysql() -> Response:
    data = request.get_json(silent=True) or {}
    connection = data.get("connection")
    table = data.get("table")
    increment = data.get("increment") or None
    limit = _to_int(data.get("limit") or 200, 200)
    analysis_mode = (data.get("analysis_mode") or "preprocess").lower()

    if not connection or not table:
        return jsonify({"status": "error", "message": "连接串和数据表名称不能为空"}), 400
    if not ALLOWED_IDENTIFIER.match(table):
        return jsonify({"status": "error", "message": "仅允许字母、数字、下划线作为表名"}), 400
    if increment and not ALLOWED_IDENTIFIER.match(increment):
        return jsonify({"status": "error", "message": "增量字段仅允许字母、数字、下划线"}), 400

    mapping = _clean_mapping(data.get("mapping"))

    try:
        params = _parse_mysql_connection(connection)
        rows_raw = _fetch_mysql_rows(params, table, increment, min(limit, 1000))
    except (RuntimeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    if not rows_raw:
        return jsonify({"status": "error", "message": "未在数据库中读取到有效数据"}), 400

    headers = list(rows_raw[0].keys())
    if analysis_mode == "deepseek":
        try:
            ai_mapping = infer_mapping_from_sample(headers, rows_raw[:5], logger=app.logger)
            mapping = _merge_mapping(mapping, ai_mapping)
        except DeepSeekError as exc:
            app.logger.warning("DeepSeek 映射失败（MySQL）：%s", exc)

    normalized_rows = _standardize_dataset(rows_raw, mapping, fallback_module=table)
    count = store.add_records(normalized_rows, source=f"mysql:{table}")
    return jsonify({"status": "success", "message": f"MySQL 导入 {count} 条", "payload": _current_payload()})


@app.post("/api/users")
def add_user_remote() -> Response:
    data = request.get_json(silent=True) or {}
    name = str(data.get("name") or "").strip()
    role = str(data.get("role") or "").strip()
    email = str(data.get("email") or "").strip()
    status = str(data.get("status") or "").strip() or "启用"

    if not name or not email:
        return jsonify({"status": "error", "message": "姓名和邮箱不能为空"}), 400

    profile = {
        "name": name,
        "role": role or "访客",
        "email": email,
        "status": status,
    }
    try:
        remote = _persist_user_to_worker(profile)
    except RuntimeError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 502

    return jsonify(
        {
            "status": "success",
            "message": "已同步到云端",
            "profile": profile,
            "remote": remote.get("data"),
        }
    )


@app.get("/api/users")
def list_users_remote() -> Response:
    try:
        users = _fetch_users_from_worker()
    except RuntimeError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 502
    return jsonify({"status": "success", "data": users})


@app.get("/docs/<path:doc_path>")
def serve_docs(doc_path: str):
    docs_dir = BASE_DIR / "docs"
    return send_from_directory(docs_dir, doc_path)


@app.get("/healthz")
def health_check() -> str:
    return "ok"


def _launch_browser() -> None:
    if os.environ.get("FLASK_BROWSER_OPENED") == "1":
        return
    url = "http://127.0.0.1:5000/"
    threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    os.environ["FLASK_BROWSER_OPENED"] = "1"


if __name__ == "__main__":
    if not os.environ.get("FLASK_SKIP_BROWSER"):
        should_open = (not DEBUG_MODE) or os.environ.get("WERKZEUG_RUN_MAIN", "").lower() == "true"
        if should_open:
            _launch_browser()
    app.run(debug=DEBUG_MODE, host="127.0.0.1", port=5000)
